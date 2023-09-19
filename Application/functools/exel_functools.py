import openpyxl as xl

import Application.functools.eval_functools as evf

from PyQt5.QtWidgets import QMessageBox
from Application.App_UI import Ui_MainWindow


def exel_lab_gkh_eval(absoluteFilePath, measure):
    wb = xl.load_workbook(rf"{absoluteFilePath}", data_only=True)
    fileName = absoluteFilePath[absoluteFilePath.rfind("\\")+1:absoluteFilePath.find(".xl")]
    ws = wb[wb.sheetnames[0]]
    resultName = None
    if measure == "no":
        resultName = "Итог в мкгСО₂/(г*ч)"
    elif measure == "gCO2perM2H":
        resultName = "Итог в гСО₂/(м²*ч)"
    elif measure == "mcgCO2perM2H":
        resultName = "Итог в мкгСО₂/(м²*ч)"
    ws["I1"] = resultName
    position = 2
    rows = ws.iter_rows(min_row=2, max_row=None, min_col=1, max_col=8)
    for x, o, b1, t, d, m, b2, e in rows:
        ra = evf.lab_gkh(x.value, o.value, b1.value, t.value, d.value, m.value, b2.value, e.value)
        if measure == "no":
            ws[f"I{position}"] = ra
        elif measure == "gCO2perM2H":
            ws[f"I{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "micro")
        elif measure == "mcgCO2perM2H":
            ws[f"I{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "no")
        position += 1
    newAbsoluteFilePath = absoluteFilePath[:absoluteFilePath.rfind("\\")+1] + fileName + "_new.xlsx"
    wb.save(rf"{newAbsoluteFilePath}")
    QMessageBox.information(Ui_MainWindow(), "Уведомление", "Новый файл будет создан в директории"
                                                              " загруженного файла. Вы можете закрыть это окно")
    del wb


def exel_lab_titr_eval(absoluteFilePath, measure):
    wb = xl.load_workbook(rf"{absoluteFilePath}", data_only=True)
    fileName = absoluteFilePath[absoluteFilePath.rfind("\\")+1:absoluteFilePath.find(".xl")]
    ws = wb[wb.sheetnames[0]]
    resultName = None
    if measure == "no":
        resultName = "Итог в мгСО₂/(г*ч)"
    elif measure == "gCO2perM2H":
        resultName = "Итог в гСО₂/(м²*ч)"
    elif measure == "mcgCO2perGH":
        resultName = "Итог в мкгСО₂/(г*ч)"
    elif measure == "mcgCO2perM2H":
        resultName = "Итог в мкгСО₂/(м²*ч)"
    ws["E1"] = resultName
    position = 2
    rows = ws.iter_rows(min_row=2, max_row=None, min_col=1, max_col=4)
    for x, o, m, e in rows:
        ra = evf.lab_titr(x.value, o.value, m.value, e.value)
        if measure == "no":
            ws[f"E{position}"] = ra
        elif measure == "mcgCO2perGH":
            ws[f"E{position}"] = ra * 10**3
        elif measure == "gCO2perM2H":
            ws[f"E{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "milli")
        elif measure == "mcgCO2perM2H":
            ws[f"E{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "milli") * 10**6
        position += 1
    newAbsoluteFilePath = absoluteFilePath[:absoluteFilePath.rfind("\\") + 1] + fileName + "_new.xlsx"
    wb.save(rf"{newAbsoluteFilePath}")
    QMessageBox.information(Ui_MainWindow(), "Уведомление", "Новый файл будет создан в директории"
                                                              " загруженного файла. Вы можете закрыть это окно")
    del wb


def exel_field_co2_eval(absoluteFilePath, measure):
    wb = xl.load_workbook(rf"{absoluteFilePath}", data_only=True)
    fileName = absoluteFilePath[absoluteFilePath.rfind("\\") + 1:absoluteFilePath.find(".xl")]
    ws = wb[wb.sheetnames[0]]
    resultName = None
    if measure == "no":
        resultName = "Итог в гСО₂/(м²*ч)"
    elif measure == "mcgCO2perGH":
        resultName = "Итог в мгСО₂/(м²*ч)"
    elif measure == "mcgCO2perGH":
        resultName = "Итог в мкгСО₂/(г*ч)"
    elif measure == "gCO2perGH":
        resultName = "Итог в гСО₂/(г*ч)"
    elif measure == "mcgCO2perM2H":
        resultName = "Итог в мкгСО₂/(м²*ч)"
    ws["G1"] = resultName
    position = 2
    rows = ws.iter_rows(min_row=2, max_row=None, min_col=1, max_col=6)
    for x, o, h, d, t, e in rows:
        ra = evf.field_co2(x.value, o.value, h.value, d.value, t.value, e.value)
        if measure == "no":
            ws[f"G{position}"] = ra
        elif measure == "mcgCO2perGH":
            ws[f"G{position}"] = evf.converter_from_GPerM2H_to_GPerGH(ra, "micro")
        elif measure == "gCO2perGH":
            ws[f"G{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "no")
        elif measure == "mgCO2perGH":
            ws[f"G{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "milli")
        elif measure == "mcgCO2perM2H":
            ws[f"G{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "mcgPerM")
        position += 1
    newAbsoluteFilePath = absoluteFilePath[:absoluteFilePath.rfind("\\") + 1] + fileName + "_new.xlsx"
    wb.save(rf"{newAbsoluteFilePath}")
    QMessageBox.information(Ui_MainWindow(), "Уведомление", "Новый файл будет создан в директории"
                                                            " загруженного файла. Вы можете закрыть это окно")
    del wb


def exel_field_gkh_eval(absoluteFilePath, measure):
    wb = xl.load_workbook(rf"{absoluteFilePath}", data_only=True)
    fileName = absoluteFilePath[absoluteFilePath.rfind("\\") + 1:absoluteFilePath.find(".xl")]
    ws = wb[wb.sheetnames[0]]
    resultName = None
    if measure == "no":
        resultName = "Итог в гСО₂/(м²*ч)"
    elif measure == "mcgCO2perGH":
        resultName = "Итог в мгСО₂/(м²*ч)"
    elif measure == "mcgCO2perGH":
        resultName = "Итог в мкгСО₂/(г*ч)"
    elif measure == "gCO2perGH":
        resultName = "Итог в гСО₂/(г*ч)"
    elif measure == "mcgCO2perM2H":
        resultName = "Итог в мкгСО₂/(м²*ч)"
    ws["H1"] = resultName
    position = 2
    rows = ws.iter_rows(min_row=2, max_row=None, min_col=1, max_col=7)
    for x, o, h, l1, l2, t, e in rows:
        ra = evf.field_gkh(x.value, o.value, h.value, l1.value, l2.value, t.value, e.value)
        if measure == "no":
            ws[f"H{position}"] = ra
        elif measure == "mcgCO2perGH":
            ws[f"H{position}"] = evf.converter_from_GPerM2H_to_GPerGH(ra, "micro")
        elif measure == "gCO2perGH":
            ws[f"H{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "no")
        elif measure == "mgCO2perGH":
            ws[f"H{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "milli")
        elif measure == "mcgCO2perM2H":
            ws[f"H{position}"] = evf.converter_from_GPerGH_to_GPerM2H(ra, "mcgPerM")
        position += 1
    newAbsoluteFilePath = absoluteFilePath[:absoluteFilePath.rfind("\\") + 1] + fileName + "_new.xlsx"
    wb.save(rf"{newAbsoluteFilePath}")
    QMessageBox.information(Ui_MainWindow(), "Уведомление", "Новый файл будет создан в директории"
                                                            " загруженного файла. Вы можете закрыть это окно")
    del wb
